#!/usr/bin/env python3
"""
Analyse automatique de fichiers Excel de trames LoRaWAN.

Pour chaque fichier (traité séparément) :
  1. Suppression des doublons de DevEUI (on garde la trame la plus récente).
     Aucune trame n'est écartée sur le SF : il ne sert qu'au niveau de pertinence.
  2. Classification de chaque capteur selon sa pertinence :
       - Indispensable   : Redondance = 1 ET SF = 7
       - Pertinence +++  : Redondance = 2 et SF dans {7, 8, 9}
       - Pertinence ++   : Redondance = 2 et SF > 9
       - Pertinence +    : Redondance 3 ou 4 (le SF n'est pas pris en compte)
       - Non pertinent   : Redondance > 5
       - À définir       : tout ce qui ne rentre dans aucune règle
                           (ex. Redondance 1 avec SF != 7, ou Redondance 5)
Dès que plusieurs fichiers sont traités ensemble (les antennes d'un même
contrat), la pertinence est calculée par recoupement entre antennes :
1 fichier = 1 antenne, et pour chaque DevEUI de chaque antenne on cherche
ce DevEUI dans les AUTRES antennes du contrat. La redondance réelle est le
nombre d'antennes du contrat qui reçoivent le capteur : s'il n'apparaît
dans aucune autre antenne ET qu'il est en SF7, c'est un très bon signal
-> Indispensable (Note 1), même si la colonne Redondance du fichier est > 1
(elle compte seulement le nombre de fois où l'antenne l'a entendu).
Cette pertinence recalculée figure dans le rapport de chaque antenne
(avec les colonnes "Nb antennes" et "Vue aussi par"), et un rapport global
du contrat est produit en plus. Avec un seul fichier, la colonne Redondance
du fichier est utilisée telle quelle.

  3. Export d'un fichier Excel de résultats :
       - Feuille "Synthèse"      : période des données, chiffres clés,
                                   répartition des pertinences + graphique
       - Feuille "Statistiques"  : distribution des SF + graphique,
                                   qualité du signal (RSSI/SNR) par catégorie,
                                   activité des capteurs (nb de trames)
       - Feuille "Tous les capteurs" + une feuille par catégorie

Suivi dans le temps : chaque analyse est historisée dans des CSV de mémoire
(datée par la date des données), puis le fichier <contrat>-suivi.xlsx est
régénéré : courbe d'évolution du contrat, stats par antenne pour chaque
analyse, et capteurs ayant changé (nouveau, disparu, amélioration,
dégradation) depuis l'analyse précédente. Les anciennes analyses sont
conservées ; relancer sur les mêmes données remplace la ligne du jour.

Organisation des dossiers : un dossier par contrat (ex. 863/), contenant un
sous-dossier Annexe/ avec les fichiers Excel des antennes (à défaut, les .xlsx
à la racine du dossier du contrat sont utilisés). Tous les fichiers générés
(rapports, analyse globale, suivi, historiques) vont dans le sous-dossier
Résultat/ du contrat, créé si besoin.

Utilisation :
    python3 analyse_pertinence.py              # traite tous les contrats du dossier courant
    python3 analyse_pertinence.py 863/         # traite un contrat précis
    python3 analyse_pertinence.py 863/ 455/    # traite plusieurs contrats
    python3 analyse_pertinence.py a.xlsx b.xlsx  # fichiers passés directement
"""

import sys
import warnings
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Ordre d'affichage des catégories et couleur associée (vert -> rouge, gris = à définir)
CATEGORIES = ["Indispensable", "Pertinence +++", "Pertinence ++", "Pertinence +",
              "Non pertinent", "À définir"]
COULEURS = {
    "Indispensable": "0CA30C",
    "Pertinence +++": "FAB219",
    "Pertinence ++": "EC835A",
    "Pertinence +": "D03B3B",
    "Non pertinent": "8B1A1A",
    "À définir": "8C8C8C",
}
# Note chiffrée associée (1 = très bon signal)
NOTES = {
    "Indispensable": 1,
    "Pertinence +++": 2,
    "Pertinence ++": 3,
    "Pertinence +": 4,
    "Non pertinent": 5,
    "À définir": None,
}
# Noms de feuille Excel (pas de '+' ni plus de 31 caractères)
NOMS_FEUILLE = {
    "Indispensable": "Indispensable",
    "Pertinence +++": "Pertinence plus-plus-plus",
    "Pertinence ++": "Pertinence plus-plus",
    "Pertinence +": "Pertinence plus",
    "Non pertinent": "Non pertinent",
    "À définir": "À définir",
}

GRAS = Font(bold=True)
TITRE = Font(bold=True, size=14)


def classer(redondance, sf) -> str:
    """Applique les règles de pertinence à un capteur.

    redondance = nombre d'antennes du contrat qui entendent le capteur
    (ou la colonne Redondance du fichier s'il n'y a qu'un seul fichier).
    """
    if redondance > 5:
        return "Non pertinent"
    if redondance == 1 and sf == 7:
        return "Indispensable"
    if redondance == 2 and sf in (7, 8, 9):
        return "Pertinence +++"
    if redondance == 2 and sf > 9:
        return "Pertinence ++"
    if redondance in (3, 4):
        return "Pertinence +"
    return "À définir"  # ex. redondance 1 avec SF != 7, ou redondance 5


def analyser_fichier(chemin: Path) -> pd.DataFrame:
    """Lit un fichier d'antenne et dédoublonne les DevEUI.

    Aucune trame n'est écartée (le SF ne sert qu'au niveau de pertinence).
    Un capteur ne compte qu'une fois par antenne : on garde sa trame la
    plus récente, et son nombre total de trames est conservé dans 'Nb trames'.
    """
    df = pd.read_excel(chemin)

    colonnes_requises = {"DevEUI", "Redondance", "SF"}
    manquantes = colonnes_requises - set(df.columns)
    if manquantes:
        raise ValueError(f"Colonnes manquantes dans {chemin.name} : {', '.join(sorted(manquantes))}")

    nb_lues = len(df)
    if "Heure" in df.columns:
        df = df.sort_values("Heure")

    # Nombre de trames émises par capteur (avant dédoublonnage) : mesure d'activité
    nb_trames = df.groupby("DevEUI").size().rename("Nb trames")

    # Dédoublonnage : une seule ligne par DevEUI, on garde la trame la plus récente
    df = df.drop_duplicates(subset="DevEUI", keep="last").reset_index(drop=True)
    df = df.merge(nb_trames, on="DevEUI")

    print(f"  {nb_lues} trames -> {len(df)} DevEUI uniques ({nb_lues - len(df)} doublons supprimés)")
    return df


def ecrire_tableau(ws, ligne, titre, table: pd.DataFrame) -> int:
    """Écrit un titre + un DataFrame dans la feuille à partir de `ligne` (1-indexé).

    Retourne la première ligne libre après le tableau.
    """
    ws.cell(row=ligne, column=1, value=titre).font = GRAS
    ligne += 1
    for j, col in enumerate(table.columns, start=1):
        ws.cell(row=ligne, column=j, value=col).font = GRAS
    for i, (_, valeurs) in enumerate(table.iterrows(), start=1):
        for j, v in enumerate(valeurs, start=1):
            ws.cell(row=ligne + i, column=j, value=v)
    return ligne + len(table) + 2


def feuille_synthese(wb, df: pd.DataFrame, meta: dict) -> None:
    ws = wb.create_sheet("Synthèse", 0)

    ws["A1"] = "Analyse de pertinence des capteurs LoRaWAN"
    ws["A1"].font = TITRE

    # Bloc d'informations générales (dont la période des données)
    ligne = 3
    for cle, valeur in meta.items():
        ws.cell(row=ligne, column=1, value=cle).font = GRAS
        ws.cell(row=ligne, column=2, value=valeur)
        ligne += 1

    # Tableau de répartition des pertinences
    compte = df["Pertinence"].value_counts().reindex(CATEGORIES, fill_value=0)
    table = pd.DataFrame({
        "Pertinence": compte.index,
        "Nombre de capteurs": compte.values,
        "%": (compte.values / len(df) * 100).round(1),
    })
    debut_table = ligne + 1
    ecrire_tableau(ws, debut_table, "Répartition par pertinence", table)

    # Graphique en secteurs de la répartition
    pie = PieChart()
    pie.title = "Répartition des capteurs par pertinence"
    data = Reference(ws, min_col=2, min_row=debut_table + 1, max_row=debut_table + 1 + len(table))
    labels = Reference(ws, min_col=1, min_row=debut_table + 2, max_row=debut_table + 1 + len(table))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    serie = pie.series[0]
    for i, cat in enumerate(CATEGORIES):
        point = DataPoint(idx=i)
        point.graphicalProperties.solidFill = COULEURS[cat]
        serie.data_points.append(point)
    pie.height, pie.width = 9, 13
    ws.add_chart(pie, "E3")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 8


def feuille_statistiques(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Statistiques")

    # Distribution des SF
    sf = df["SF"].value_counts().sort_index()
    table_sf = pd.DataFrame({"SF": sf.index, "Nombre de capteurs": sf.values})
    ligne = ecrire_tableau(ws, 1, "Distribution des SF (Spreading Factor)", table_sf)

    bar = BarChart()
    bar.type = "col"
    bar.title = "Nombre de capteurs par SF"
    bar.legend = None
    data = Reference(ws, min_col=2, min_row=2, max_row=2 + len(table_sf))
    labels = Reference(ws, min_col=1, min_row=3, max_row=2 + len(table_sf))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.series[0].graphicalProperties.solidFill = "4472C4"
    bar.height, bar.width = 8, 12
    ws.add_chart(bar, "E1")

    # Qualité du signal par catégorie de pertinence
    if {"RSSI", "SNR"} <= set(df.columns):
        stats = (
            df.groupby("Pertinence")
            .agg(**{
                "Nb capteurs": ("DevEUI", "count"),
                "RSSI moyen (dBm)": ("RSSI", "mean"),
                "RSSI min": ("RSSI", "min"),
                "RSSI max": ("RSSI", "max"),
                "SNR moyen (dB)": ("SNR", "mean"),
            })
            .reindex([c for c in CATEGORIES if c in df["Pertinence"].values])
            .round(1)
            .reset_index()
        )
        ligne = ecrire_tableau(ws, max(ligne, 18), "Qualité du signal par pertinence", stats)

    # Activité des capteurs (nombre de trames émises sur la période)
    actifs = df.nlargest(10, "Nb trames")[["DevEUI", "Nb trames", "SF", "Pertinence"]]
    ws.cell(row=ligne, column=1,
            value=f"Trames par capteur : moyenne {df['Nb trames'].mean():.1f}, "
                  f"médiane {df['Nb trames'].median():.0f}, max {df['Nb trames'].max()}").font = GRAS
    ecrire_tableau(ws, ligne + 1, "Top 10 des capteurs les plus actifs", actifs)

    for col, largeur in zip("ABCDEF", (26, 18, 12, 12, 12, 16)):
        ws.column_dimensions[col].width = largeur


def exporter_resultats(df: pd.DataFrame, meta: dict, sortie: Path) -> None:
    """Écrit le fichier Excel de résultats : synthèse, statistiques et détail."""
    with pd.ExcelWriter(sortie, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Tous les capteurs", index=False)
        for cat in CATEGORIES:
            sous_df = df[df["Pertinence"] == cat]
            if not sous_df.empty:
                sous_df.to_excel(writer, sheet_name=NOMS_FEUILLE[cat], index=False)

        feuille_synthese(writer.book, df, meta)
        feuille_statistiques(writer.book, df)

    print(f"  Résultats écrits dans : {sortie}")
    compte = df["Pertinence"].value_counts().reindex(CATEGORIES, fill_value=0)
    for cat, nb in compte.items():
        print(f"    {cat:<15} {nb:>5}  ({nb / len(df) * 100:.1f} %)")


def analyse_globale(resultats: list) -> pd.DataFrame:
    """Analyse croisée multi-antennes : 1 fichier = 1 antenne.

    Pour chaque DevEUI, la redondance réelle est le nombre d'antennes
    différentes qui le reçoivent. Un capteur vu par une seule antenne
    est indispensable. Le SF retenu est le meilleur (le plus bas)
    observé parmi les antennes.
    """
    trames = []
    for chemin, df in resultats:
        d = df.copy()
        d["Antenne"] = chemin.stem
        trames.append(d)
    tout = pd.concat(trames, ignore_index=True)

    agregats = {
        "Nb antennes": ("Antenne", "nunique"),
        "Antennes": ("Antenne", lambda s: ", ".join(sorted(set(s)))),
        "SF": ("SF", "min"),
        "Nb trames": ("Nb trames", "sum"),
    }
    if "Heure" in tout.columns:
        agregats["Dernière trame"] = ("Heure", "max")

    capteurs = tout.groupby("DevEUI").agg(**agregats).reset_index()
    capteurs["Pertinence"] = [classer(n, s) for n, s in zip(capteurs["Nb antennes"], capteurs["SF"])]
    capteurs["Note"] = [NOTES[p] for p in capteurs["Pertinence"]]
    return capteurs


def exporter_globale(resultats: list, dossier_sortie: Path, nom_contrat: str) -> pd.DataFrame:
    capteurs = analyse_globale(resultats)

    meta = {
        "Type d'analyse": "Globale multi-antennes (redondance = nb d'antennes recevant le capteur)",
        "Date de l'analyse": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "Antennes analysées": ", ".join(chemin.stem for chemin, _ in resultats),
        "Nombre d'antennes": len(resultats),
        "Capteurs uniques (DevEUI)": len(capteurs),
        "Capteurs vus par une seule antenne": int((capteurs["Nb antennes"] == 1).sum()),
    }
    if "Dernière trame" in capteurs.columns:
        heures = pd.concat([df["Heure"] for _, df in resultats if "Heure" in df.columns])
        meta["Période des données"] = f"du {heures.min():%d/%m/%Y %H:%M} au {heures.max():%d/%m/%Y %H:%M}"
        suffixe = f"-{heures.max():%d%m}"
    else:
        suffixe = ""

    sortie = dossier_sortie / f"{nom_contrat}-analyse-globale{suffixe}.xlsx"
    print(f"\nAnalyse globale du contrat '{nom_contrat}' ({len(resultats)} antennes)")
    exporter_resultats(capteurs, meta, sortie)
    return capteurs


def actualiser_csv(chemin: Path, nouvelles_lignes: pd.DataFrame, date_str: str) -> pd.DataFrame:
    """Ajoute les lignes du jour à un CSV d'historique (remplace la même date)."""
    if chemin.exists():
        historique = pd.read_csv(chemin)
        historique = historique[historique["Date"] != date_str]
        historique = pd.concat([historique, nouvelles_lignes], ignore_index=True)
    else:
        historique = nouvelles_lignes
    historique = historique.sort_values("Date", kind="stable").reset_index(drop=True)
    historique.to_csv(chemin, index=False)
    return historique


def compter_pertinences(df: pd.DataFrame) -> dict:
    compte = df["Pertinence"].value_counts().reindex(CATEGORIES, fill_value=0)
    return {c: int(n) for c, n in compte.items()}


def historiser_analyse(antennes: list, capteurs: pd.DataFrame, dossier: Path,
                       nom: str, date_donnees) -> None:
    """Historise l'analyse dans les CSV de mémoire du contrat.

    L'analyse est datée par la date des données (la plus récente trouvée).
    Trois fichiers CSV servent de mémoire entre les exécutions :
      - <contrat>-historique-syntheses.csv : une ligne par analyse (contrat entier)
      - <contrat>-historique-antennes.csv  : une ligne par antenne et par analyse
      - <contrat>-historique-capteurs.csv  : la pertinence de chaque DevEUI par analyse
    Relancer sur les mêmes données (même date) remplace les lignes (pas de doublon).
    """
    if date_donnees is not None and not pd.isna(date_donnees):
        date_str = f"{pd.Timestamp(date_donnees):%Y-%m-%d}"
    else:
        date_str = datetime.now().strftime("%Y-%m-%d")

    ligne = {"Date": date_str, "Capteurs": len(capteurs), **compter_pertinences(capteurs)}
    actualiser_csv(dossier / f"{nom}-historique-syntheses.csv", pd.DataFrame([ligne]), date_str)

    lignes_antennes = [
        {"Date": date_str, "Antenne": fichier.stem, "Capteurs": len(df), **compter_pertinences(df)}
        for fichier, df in antennes
    ]
    actualiser_csv(dossier / f"{nom}-historique-antennes.csv", pd.DataFrame(lignes_antennes), date_str)

    detail = capteurs[["DevEUI", "Pertinence", "Note"]].copy()
    detail.insert(0, "Date", date_str)
    actualiser_csv(dossier / f"{nom}-historique-capteurs.csv", detail, date_str)


def regenerer_suivi(dossier: Path, nom: str) -> None:
    """Régénère <contrat>-suivi.xlsx à partir des CSV d'historique :
    courbe d'évolution du contrat, stats par antenne pour chaque journée
    analysée, et capteurs ayant changé depuis la journée précédente.
    """
    from openpyxl import Workbook
    from openpyxl.chart import LineChart

    historique = pd.read_csv(dossier / f"{nom}-historique-syntheses.csv")
    hist_antennes = pd.read_csv(dossier / f"{nom}-historique-antennes.csv")
    hist_capteurs = pd.read_csv(dossier / f"{nom}-historique-capteurs.csv")

    wb = Workbook()

    ws = wb.active
    ws.title = "Évolution"
    ws["A1"] = f"Suivi de la pertinence — contrat {nom}"
    ws["A1"].font = TITRE
    ecrire_tableau(ws, 3, "Historique des journées analysées (contrat entier)", historique)

    graph = LineChart()
    graph.title = "Évolution du nombre de capteurs par pertinence"
    graph.y_axis.title = "Nombre de capteurs"
    # Colonnes : A=Date, B=Capteurs, C.. = catégories (lignes 4=en-têtes, 5..=données)
    data = Reference(ws, min_col=3, max_col=2 + len(CATEGORIES), min_row=4, max_row=4 + len(historique))
    graph.add_data(data, titles_from_data=True)
    graph.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4 + len(historique)))
    for i, cat in enumerate(CATEGORIES):
        serie = graph.series[i]
        serie.graphicalProperties.line.solidFill = COULEURS[cat]
        serie.graphicalProperties.line.width = 25000  # ~2 pt
        serie.smooth = False
    graph.height, graph.width = 10, 22
    ws.add_chart(graph, f"A{6 + len(historique)}")
    ws.column_dimensions["A"].width = 14

    # Feuille "Par antenne" : les stats de chaque antenne du contrat,
    # une ligne par date d'analyse
    ws_ant = wb.create_sheet("Par antenne")
    ws_ant["A1"] = f"Suivi par antenne — contrat {nom}"
    ws_ant["A1"].font = TITRE
    ligne_courante = 3
    for antenne in sorted(hist_antennes["Antenne"].unique()):
        bloc = hist_antennes[hist_antennes["Antenne"] == antenne].drop(columns="Antenne")
        ligne_courante = ecrire_tableau(ws_ant, ligne_courante, f"Antenne : {antenne}", bloc)
    ws_ant.column_dimensions["A"].width = 14
    for col in "BCDEFGH":
        ws_ant.column_dimensions[col].width = 15

    ws2 = wb.create_sheet("Changements")
    dates = sorted(hist_capteurs["Date"].unique())
    if len(dates) < 2:
        ws2["A1"] = "Première analyse : les changements apparaîtront à partir de la prochaine exécution."
    else:
        avant, apres = dates[-2], dates[-1]
        p_avant = hist_capteurs[hist_capteurs["Date"] == avant].set_index("DevEUI")["Pertinence"]
        p_apres = hist_capteurs[hist_capteurs["Date"] == apres].set_index("DevEUI")["Pertinence"]
        lignes = []
        for eui in sorted(set(p_avant.index) | set(p_apres.index)):
            av, ap = p_avant.get(eui), p_apres.get(eui)
            if av is None:
                lignes.append({"DevEUI": eui, "Avant": "—", "Après": ap, "Changement": "Nouveau capteur"})
            elif ap is None:
                lignes.append({"DevEUI": eui, "Avant": av, "Après": "—", "Changement": "Capteur disparu"})
            elif av != ap:
                sens = "Amélioration" if (NOTES.get(ap) or 9) < (NOTES.get(av) or 9) else "Dégradation"
                lignes.append({"DevEUI": eui, "Avant": av, "Après": ap, "Changement": sens})
        ws2["A1"] = f"Changements entre le {avant} et le {apres}"
        ws2["A1"].font = GRAS
        if lignes:
            ecrire_tableau(ws2, 3, f"{len(lignes)} capteur(s) concerné(s)", pd.DataFrame(lignes))
        else:
            ws2["A3"] = "Aucun changement de pertinence."
        for col, largeur in zip("ABCD", (26, 16, 16, 20)):
            ws2.column_dimensions[col].width = largeur

    sortie = dossier / f"{nom}-suivi.xlsx"
    wb.save(sortie)
    print(f"\nSuivi mis à jour : {sortie} ({len(historique)} journée(s) dans l'historique)")


def filtrer_sorties(fichiers) -> list:
    """Écarte les fichiers générés par le script lui-même."""
    return [f for f in sorted(fichiers)
            if not f.stem.endswith(("-analyse", "_resultats", "-suivi"))
            and "-analyse-globale" not in f.stem
            and not f.name.startswith("~$")]


def fichiers_du_contrat(dossier: Path) -> list:
    """Fichiers des antennes d'un contrat : dans Annexe/ si présent, sinon à la racine."""
    for sous in dossier.iterdir():
        if sous.is_dir() and sous.name.lower() in ("annexe", "annexes"):
            xlsx = filtrer_sorties(sous.glob("*.xlsx"))
            if xlsx:
                return xlsx
    return filtrer_sorties(dossier.glob("*.xlsx"))


def dossier_resultat(dossier: Path) -> Path:
    """Renvoie le sous-dossier Résultat du contrat (créé si absent)."""
    for sous in dossier.iterdir():
        if sous.is_dir() and sous.name.lower() in ("résultat", "resultat", "résultats", "resultats"):
            return sous
    sortie = dossier / "Résultat"
    sortie.mkdir(exist_ok=True)
    return sortie


def croiser_antennes(antennes: list) -> tuple:
    """Classe les capteurs du contrat. antennes = [(fichier, df), ...]

    Avec plusieurs antennes : pour chaque DevEUI de chaque antenne, on cherche
    ce DevEUI dans les autres antennes du contrat. Un capteur entendu par une
    seule antenne du contrat, en SF7, est Indispensable (Note 1). La colonne
    Redondance du fichier n'est pas utilisée : elle compte le nombre de fois
    où l'antenne a entendu le capteur, pas le nombre d'antennes. Avec une
    seule antenne, la colonne Redondance est utilisée telle quelle.

    Retourne (liste (fichier, df) classée, df global des capteurs du contrat).
    """
    if len(antennes) >= 2:
        vu_par: dict = {}
        for fichier, df in antennes:
            for eui in df["DevEUI"]:
                vu_par.setdefault(eui, set()).add(fichier.stem)

        maj = []
        for fichier, df in antennes:
            df = df.copy()
            autres = [", ".join(sorted(vu_par[e] - {fichier.stem})) for e in df["DevEUI"]]
            df["Nb antennes"] = [len(vu_par[e]) for e in df["DevEUI"]]
            df["Vue aussi par"] = [a if a else "Aucune autre antenne" for a in autres]
            df["Pertinence"] = [classer(n, s) for n, s in zip(df["Nb antennes"], df["SF"])]
            df["Note"] = [NOTES[p] for p in df["Pertinence"]]
            df = df[[c for c in df.columns if c not in ("Pertinence", "Note")] + ["Pertinence", "Note"]]
            maj.append((fichier, df))
        capteurs = analyse_globale(maj)
    else:
        fichier, df = antennes[0]
        df = df.copy()
        df["Pertinence"] = [classer(r, s) for r, s in zip(df["Redondance"], df["SF"])]
        df["Note"] = [NOTES[p] for p in df["Pertinence"]]
        maj = [(fichier, df)]
        capteurs = df
    return maj, capteurs


def traiter_contrat(nom: str, fichiers: list, dossier: Path) -> None:
    """Pipeline complet d'un contrat : rapports par antenne, analyse globale
    et mise à jour du suivi. Tout va dans le sous-dossier Résultat/ du contrat.
    """
    sortie_dir = dossier_resultat(dossier)

    # Lecture et dédoublonnage de chaque antenne
    bruts = []
    for fichier in fichiers:
        print(f"\nLecture de : {fichier.name}")
        try:
            bruts.append((fichier, analyser_fichier(fichier)))
        except Exception as e:
            print(f"  ERREUR : {e}")
    if not bruts:
        return

    antennes, capteurs = croiser_antennes(bruts)

    heures = [df["Heure"].max() for _, df in antennes
              if "Heure" in df.columns and df["Heure"].notna().any()]
    date_donnees = max(heures) if heures else None
    suffixe = f"-{date_donnees:%d%m}" if date_donnees is not None else ""

    # Rapport de chaque antenne dans Résultat/
    for fichier, df in antennes:
        meta = {
            "Fichier source": fichier.name,
            "Date de l'analyse": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "Capteurs uniques (DevEUI)": len(df),
        }
        if "Heure" in df.columns and df["Heure"].notna().any():
            meta["Période des données"] = (f"du {df['Heure'].min():%d/%m/%Y %H:%M} "
                                           f"au {df['Heure'].max():%d/%m/%Y %H:%M}")
        if len(antennes) >= 2:
            meta["Antennes du contrat"] = ", ".join(f.stem for f, _ in antennes)
            meta["Capteurs vus uniquement par cette antenne"] = int((df["Nb antennes"] == 1).sum())
        print(f"\nRapport de l'antenne : {fichier.stem}")
        exporter_resultats(df, meta, sortie_dir / f"{fichier.stem}{suffixe}-analyse.xlsx")

    # Analyse globale du contrat, puis historisation et suivi
    if len(antennes) >= 2:
        exporter_globale(antennes, sortie_dir, nom)
    historiser_analyse(antennes, capteurs, sortie_dir, nom, date_donnees)
    regenerer_suivi(sortie_dir, nom)


def main() -> None:
    args = sys.argv[1:]
    cibles = [Path(a) for a in args] if args else [Path(".")]

    # Découverte des contrats à traiter :
    #  - les sous-dossiers contenant des .xlsx (directement ou dans Annexe/)
    #    sont les contrats ; les .xlsx isolés à la racine sont alors ignorés
    #  - sinon, le dossier lui-même est un contrat s'il contient des .xlsx
    #  - des fichiers .xlsx passés directement = un contrat ad hoc
    dossiers_reserves = ("annexe", "annexes", "résultat", "resultat", "résultats", "resultats")
    contrats: list = []
    fichiers_directs: list = []
    for cible in cibles:
        if cible.is_file():
            fichiers_directs.append(cible)
        elif cible.is_dir():
            sous_contrats = []
            for sous in sorted(cible.iterdir()):
                if (sous.is_dir() and not sous.name.startswith((".", "_"))
                        and sous.name.lower() not in dossiers_reserves):
                    xlsx = fichiers_du_contrat(sous)
                    if xlsx:
                        sous_contrats.append((sous.name, xlsx, sous))
            if sous_contrats:
                contrats.extend(sous_contrats)
                if filtrer_sorties(cible.glob("*.xlsx")):
                    print(f"Info : .xlsx à la racine de {cible.resolve().name} ignorés "
                          f"(les contrats sont les sous-dossiers)")
            else:
                xlsx = fichiers_du_contrat(cible)
                if xlsx:
                    contrats.append((cible.resolve().name or "contrat", xlsx, cible))
        else:
            print(f"Ignoré (introuvable) : {cible}")

    if fichiers_directs:
        fichiers_directs = filtrer_sorties(fichiers_directs)
        if fichiers_directs:
            parent = fichiers_directs[0].parent
            contrats.append((parent.resolve().name or "contrat", fichiers_directs, parent))

    if not contrats:
        print("Aucun fichier .xlsx à traiter.")
        sys.exit(1)

    for nom, fichiers, dossier in contrats:
        print(f"\n{'=' * 50}\nContrat : {nom} ({len(fichiers)} fichier(s) d'antenne)\n{'=' * 50}")
        traiter_contrat(nom, fichiers, dossier)


if __name__ == "__main__":
    main()
